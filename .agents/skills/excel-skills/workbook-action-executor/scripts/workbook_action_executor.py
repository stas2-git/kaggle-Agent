#!/usr/bin/env python3
"""Deterministic workbook actions for repeatable Excel edits."""

from __future__ import annotations

import argparse
import json
import subprocess
import sys
from copy import copy
from datetime import datetime, timezone
from pathlib import Path
from types import SimpleNamespace
from typing import Any

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from _shared.llm_work_audit import (  # type: ignore  # noqa: E402
    append_run_event,
    artifact_path,
    file_metadata,
    infer_llm_work_root,
    task_fingerprint,
    timestamp_run_id,
    workbook_metadata,
)

NUMBER_FORMAT_PRESETS = {
    "accounting": '_(* #,##0_);_(* \\(#,##0\\);_(* "-"_);_(@_)',
    "accounting_usd": '_($* #,##0_);_($* \\(#,##0\\);_($* "-"_);_(@_)',
    "percentage": "0.00%",
    "percentage_whole": "0%",
    "currency": "$#,##0.00",
    "currency_whole": "$#,##0",
    "date_short": "m/d/yy",
    "date_long": "mm/dd/yyyy",
    "number_comma": "#,##0",
    "number_comma_2dec": "#,##0.00",
    "number_2dec": "0.00",
}

DEFAULT_THEME = {
    "accent_fill": "1F4E78",
    "accent_fill_light": "D9EAF7",
    "header_font_color": "FFFFFF",
    "border_color": "9FBAD0",
    "input_fill": "FFF2CC",
    "formula_fill": "E2F0D9",
    "total_fill": "1F4E78",
    "note_fill": "F3F4F6",
    "neutral_fill": "F8FBFD",
    "neutral_text": "44546A",
}

STYLE_PRESETS = {
    "section_header": {
        "font": {"bold": True, "size": 14, "color": "{header_font_color}"},
        "fill": {"color": "{accent_fill}"},
        "alignment": {"vertical": "center", "wrap_text": True},
        "border": {"style": "medium", "color": "{border_color}", "sides": ["bottom"]},
    },
    "table_header": {
        "font": {"bold": True, "size": 11, "color": "{header_font_color}"},
        "fill": {"color": "{accent_fill}"},
        "alignment": {"horizontal": "center", "vertical": "center", "wrap_text": True},
        "border": {"style": "thin", "color": "{border_color}", "sides": "all"},
    },
    "input_cell": {
        "fill": {"color": "{input_fill}"},
        "border": {"style": "thin", "color": "{border_color}", "sides": "all"},
    },
    "formula_cell": {
        "fill": {"color": "{formula_fill}"},
        "border": {"style": "thin", "color": "{border_color}", "sides": "all"},
    },
    "final_total": {
        "font": {"bold": True, "color": "{header_font_color}"},
        "fill": {"color": "{total_fill}"},
        "border": {"style": "medium", "color": "{border_color}", "sides": ["top", "bottom"]},
        "alignment": {"vertical": "center"},
    },
    "note_cell": {
        "font": {"italic": True, "color": "{neutral_text}"},
        "fill": {"color": "{note_fill}"},
        "alignment": {"wrap_text": True, "vertical": "top"},
    },
}


def resolve_path(value: str) -> Path:
    return Path(value).expanduser().resolve()


def infer_value(raw: str) -> Any:
    if not isinstance(raw, str):
        return raw
    text = raw.strip()
    if text == "":
        return ""
    if text.startswith("="):
        return text
    lowered = text.lower()
    if lowered == "true":
        return True
    if lowered == "false":
        return False
    if lowered == "none":
        return None
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
        try:
            dt = datetime.strptime(text, fmt)
            excel_epoch = datetime(1899, 12, 30)
            delta = dt - excel_epoch
            return delta.days + (delta.seconds / 86400.0)
        except ValueError:
            continue
    try:
        if "." not in text and "e" not in lowered:
            return int(text)
        return float(text)
    except ValueError:
        return raw


def parse_assignment(raw: str) -> tuple[str, str]:
    if "=" not in raw:
        raise ValueError(f"Expected assignment like A1=value, got: {raw}")
    left, right = raw.split("=", 1)
    left = left.strip()
    if not left:
        raise ValueError(f"Missing target in assignment: {raw}")
    return left, right


def normalize_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        lowered = value.strip().lower()
        if lowered in {"true", "1", "yes", "on"}:
            return True
        if lowered in {"false", "0", "no", "off"}:
            return False
    return bool(value)


def load_plan(path: Path) -> list[dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, dict):
        actions = data.get("actions")
        if not isinstance(actions, list):
            raise ValueError("Plan JSON object must contain an 'actions' list.")
        return actions
    if isinstance(data, list):
        return data
    raise ValueError("Plan must be a JSON list or an object with an 'actions' list.")


def infer_action_paths(
    workbook_path: Path, llm_work_root_value: str | None, run_id_value: str | None
) -> tuple[Path, str]:
    llm_work_root = infer_llm_work_root(workbook_path, llm_work_root_value)
    run_id = run_id_value or timestamp_run_id()
    return llm_work_root, run_id


def save_action_artifact(
    llm_work_root: Path,
    run_id: str,
    command_name: str,
    payload: dict[str, object],
    _no_mirror_current: bool,
) -> tuple[Path, Path | None]:
    output_path = artifact_path(llm_work_root, run_id, f"action-{command_name}.json")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    return output_path, None


def workbook_sheet(wb, name: str | None) -> Worksheet:
    if name:
        return wb[name]
    return wb.active


def build_workbook_state(wb, workbook_path: Path) -> dict[str, object]:
    is_read_only = bool(getattr(wb, "read_only", False))
    is_data_only = bool(getattr(wb, "data_only", False))
    return {
        "is_open": True,
        "filename": str(workbook_path),
        "active_sheet": wb.active.title if wb.active else None,
        "sheets": wb.sheetnames,
        "sheet_count": len(wb.sheetnames),
        "is_read_only": is_read_only,
        "is_data_only": is_data_only,
        "warning": (
            "Workbook is in read-only or data-only mode; saving may fail."
            if (is_read_only or is_data_only)
            else None
        ),
    }


def save_workbook_safely(wb, workbook_path: Path) -> str:
    if wb is None:
        raise ValueError("No workbook loaded.")
    if bool(getattr(wb, "read_only", False)):
        raise ValueError(
            f"Cannot save '{workbook_path.name}' because the workbook is in read-only mode."
        )
    if bool(getattr(wb, "data_only", False)):
        raise ValueError(
            f"Cannot save '{workbook_path.name}' because the workbook was loaded in data-only mode."
        )
    try:
        wb.save(workbook_path)
    except (IOError, PermissionError, ValueError, AttributeError) as exc:
        message = str(exc).lower()
        if "i/o operation" in message or "closed file" in message:
            raise ValueError(
                f"Cannot save '{workbook_path.name}' because the workbook file handle is broken. Reload the workbook fresh and rerun the action."
            ) from exc
        if "permission" in message or "read-only" in message:
            raise ValueError(
                f"Cannot save '{workbook_path.name}' because the file is locked or read-only."
            ) from exc
        raise ValueError(f"Failed to save workbook: {exc}") from exc
    return f"Saved workbook: {workbook_path}"


def cells_from_reference(ws: Worksheet, cell_reference: str) -> list[Any]:
    if ":" not in cell_reference:
        return [ws[cell_reference]]
    cells = []
    for row in ws[cell_reference]:
        if isinstance(row, tuple):
            cells.extend(row)
        else:
            cells.append(row)
    return cells


def iter_range_cells(ws: Worksheet, cell_reference: str):
    if ":" not in cell_reference:
        yield 0, 0, ws[cell_reference]
        return
    min_col, min_row, max_col, max_row = range_boundaries(cell_reference)
    for r_offset, row in enumerate(
        ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)
    ):
        for c_offset, cell in enumerate(row):
            yield r_offset, c_offset, cell


def merged_ranges_within(ws: Worksheet, cell_reference: str) -> list[str]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_reference)
    ranges: list[str] = []
    for merged in ws.merged_cells.ranges:
        if (
            merged.min_col >= min_col
            and merged.max_col <= max_col
            and merged.min_row >= min_row
            and merged.max_row <= max_row
        ):
            ranges.append(str(merged))
    return ranges


def maybe_json_or_object(value: Any) -> dict[str, Any] | None:
    if value is None:
        return None
    if isinstance(value, dict):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        loaded = json.loads(text)
        if not isinstance(loaded, dict):
            raise ValueError("Expected JSON object.")
        return loaded
    raise ValueError(f"Expected JSON object or JSON string, got {type(value).__name__}.")


def merge_dicts(base: dict[str, Any], override: dict[str, Any]) -> dict[str, Any]:
    merged: dict[str, Any] = {**base}
    for key, value in override.items():
        if isinstance(value, dict) and isinstance(merged.get(key), dict):
            merged[key] = merge_dicts(merged[key], value)
        else:
            merged[key] = value
    return merged


def resolve_theme_value(value: Any, theme: dict[str, Any]) -> Any:
    if isinstance(value, str) and value.startswith("{") and value.endswith("}"):
        return theme.get(value[1:-1], value)
    if isinstance(value, dict):
        return {k: resolve_theme_value(v, theme) for k, v in value.items()}
    if isinstance(value, list):
        return [resolve_theme_value(v, theme) for v in value]
    return value


def resolve_theme(theme_value: Any = None) -> dict[str, Any]:
    theme = dict(DEFAULT_THEME)
    theme_override = maybe_json_or_object(theme_value) if theme_value is not None else None
    if theme_override:
        theme.update(theme_override)
    return theme


def parse_range_list(single_range: Any = None, ranges_value: Any = None) -> list[str]:
    ranges: list[str] = []
    if single_range:
        if isinstance(single_range, list):
            ranges.extend(str(item) for item in single_range if item)
        else:
            ranges.append(str(single_range))
    if ranges_value:
        if isinstance(ranges_value, list):
            ranges.extend(str(item) for item in ranges_value if item)
        else:
            ranges.append(str(ranges_value))
    seen: set[str] = set()
    ordered: list[str] = []
    for value in ranges:
        if value not in seen:
            ordered.append(value)
            seen.add(value)
    return ordered


def unique_cells_for_ranges(ws: Worksheet, ranges: list[str]) -> list[Any]:
    cells = []
    seen: set[str] = set()
    for cell_reference in ranges:
        for cell in cells_from_reference(ws, cell_reference):
            if cell.coordinate not in seen:
                cells.append(cell)
                seen.add(cell.coordinate)
    return cells


def resolve_number_format(value: str | None) -> str | None:
    if not value:
        return None
    return NUMBER_FORMAT_PRESETS.get(value, value)


def normalize_color(value: str | None) -> str | None:
    if value is None:
        return None
    color = str(value).replace("#", "").upper()
    if len(color) == 6:
        return "FF" + color
    return color


def build_side(style: dict[str, Any], fallback_color: str | None = None) -> Side:
    color = normalize_color(style.get("color") or fallback_color)
    kwargs = {"style": style.get("style", "thin")}
    if color:
        kwargs["color"] = color
    return Side(**kwargs)


def build_font(style: dict[str, Any], base: Font | None = None) -> Font:
    font = copy(base) if base is not None else Font()
    kwargs = {
        "name": style.get("name", font.name),
        "size": style.get("size", font.sz),
        "bold": style.get("bold", font.bold),
        "italic": style.get("italic", font.italic),
        "underline": style.get("underline", font.underline),
        "strike": style.get("strikethrough", font.strike),
    }
    color = normalize_color(style.get("color"))
    if color is not None:
        kwargs["color"] = color
    return Font(**kwargs)


def build_fill(style: dict[str, Any]) -> PatternFill:
    color = normalize_color(style.get("fg_color") or style.get("color") or "FFFFFF") or "FFFFFFFF"
    return PatternFill(
        fill_type=style.get("pattern_type", "solid"),
        start_color=color,
        end_color=color,
    )


def build_border(style: dict[str, Any], base: Border | None = None) -> Border:
    color = style.get("color")
    default_side = build_side(style, color)
    border = copy(base) if base is not None else Border()
    sides = style.get("sides", "all")
    if isinstance(sides, str):
        if sides in {"all", "outline"}:
            wanted = {"left", "right", "top", "bottom"}
        else:
            wanted = {sides}
    else:
        wanted = set(sides)

    side_specs = {}
    for name in ("left", "right", "top", "bottom"):
        side_value = style.get(name)
        if side_value:
            side_specs[name] = build_side(side_value, color)
        elif name in wanted:
            side_specs[name] = default_side
        else:
            side_specs[name] = getattr(border, name)
    return Border(**side_specs)


def build_alignment(style: dict[str, Any], base: Alignment | None = None) -> Alignment:
    align = copy(base) if base is not None else Alignment()
    return Alignment(
        horizontal=style.get("horizontal", align.horizontal),
        vertical=style.get("vertical", align.vertical),
        wrap_text=style.get("wrap_text", align.wrap_text),
        shrink_to_fit=style.get("shrink_to_fit", align.shrink_to_fit),
        indent=style.get("indent", align.indent),
        text_rotation=style.get("text_rotation", align.text_rotation),
    )


def build_style_spec(args) -> tuple[dict[str, Any], dict[str, Any]]:
    theme = resolve_theme(getattr(args, "theme", None))
    preset_names = getattr(args, "preset", None) or []
    if isinstance(preset_names, str):
        preset_names = [preset_names]

    spec: dict[str, Any] = {}
    resolved_presets: list[str] = []
    for preset_name in preset_names:
        if preset_name not in STYLE_PRESETS:
            raise ValueError(
                f"Unknown style preset '{preset_name}'. Available presets: {', '.join(sorted(STYLE_PRESETS))}."
            )
        preset = resolve_theme_value(STYLE_PRESETS[preset_name], theme)
        spec = merge_dicts(spec, preset)
        resolved_presets.append(preset_name)

    overrides = {
        "font": resolve_theme_value(maybe_json_or_object(getattr(args, "font", None)), theme)
        if getattr(args, "font", None) is not None
        else None,
        "fill": resolve_theme_value(maybe_json_or_object(getattr(args, "fill", None)), theme)
        if getattr(args, "fill", None) is not None
        else None,
        "border": resolve_theme_value(maybe_json_or_object(getattr(args, "border", None)), theme)
        if getattr(args, "border", None) is not None
        else None,
        "alignment": resolve_theme_value(
            maybe_json_or_object(getattr(args, "alignment", None)), theme
        )
        if getattr(args, "alignment", None) is not None
        else None,
        "number_format": resolve_number_format(getattr(args, "number_format", None)),
    }
    for key, value in overrides.items():
        if value is not None:
            spec[key] = value
    return spec, {"theme": theme, "presets": resolved_presets}


def apply_style_to_cells(cells: list[Any], spec: dict[str, Any]) -> int:
    changed = 0
    font_style = spec.get("font")
    fill_style = spec.get("fill")
    border_style = spec.get("border")
    alignment_style = spec.get("alignment")
    number_format = spec.get("number_format")
    for cell in cells:
        if font_style is not None:
            cell.font = build_font(font_style, cell.font)
        if fill_style is not None:
            cell.fill = build_fill(fill_style)
        if border_style is not None:
            cell.border = build_border(border_style, cell.border)
        if alignment_style is not None:
            cell.alignment = build_alignment(alignment_style, cell.alignment)
        if number_format is not None:
            cell.number_format = number_format
        changed += 1
    return changed


def normalize_style_action_payload(step: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(step)
    style_payload = normalized.get("style")
    if isinstance(style_payload, dict):
        for key in ("font", "fill", "border", "alignment", "number_format"):
            if key in style_payload and key not in normalized:
                normalized[key] = style_payload[key]
    return normalized


def normalize_plan_action(step: dict[str, Any]) -> dict[str, Any]:
    normalized = normalize_style_action_payload(step)
    if normalized.get("range") and "ranges" not in normalized:
        normalized["ranges"] = [normalized["range"]]
    for key in ("copy_values", "copy_formulas", "copy_styles", "copy_merges", "copy_dimensions", "strict", "activate", "show_gridlines", "center_horizontally", "center_vertically", "autofit"):
        if key in normalized:
            normalized[key] = normalize_bool(normalized[key])
    return normalized


PLAN_REQUIRED_FIELDS = {
    "create-sheet": ("sheet",),
    "rename-sheet": ("sheet", "new_name"),
    "reorder-sheet": ("sheet", "index"),
    "delete-sheet": ("sheet",),
    "write-cells": ("set",),
    "write-formulas": ("set",),
    "fill-formula": ("source_cell", "target_range"),
    "copy-range": ("source", "target"),
    "define-name": ("name", "refers_to"),
    "create-table": ("name", "range"),
    "insert-rows": ("index",),
    "delete-rows": ("index",),
    "insert-cols": ("index",),
    "delete-cols": ("index",),
    "set-column-widths": ("set",),
    "set-row-heights": ("set",),
    "set-style": ("ranges",),
    "format-range": ("ranges",),
    "freeze-panes": ("cell",),
    "merge-cells": ("range",),
    "unmerge-cells": ("range",),
    "set-sheet-view": ("sheet",),
    "cleanup-sheet": ("sheet",),
}


def validate_plan_actions(actions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    for idx, original_step in enumerate(actions, start=1):
        if not isinstance(original_step, dict):
            issues.append({"step": idx, "error": "Plan step must be an object."})
            continue
        step = normalize_plan_action(original_step)
        command = step.get("command")
        if not command:
            issues.append({"step": idx, "error": "Plan step is missing 'command'."})
            continue
        if command == "run-plan":
            issues.append(
                {"step": idx, "command": command, "error": "Nested run-plan is not supported."}
            )
            continue
        if command not in COMMAND_MAP:
            issues.append({"step": idx, "command": command, "error": "Unsupported command."})
            continue

        missing = []
        for field in PLAN_REQUIRED_FIELDS.get(command, ()):
            value = step.get(field)
            if value is None or value == [] or value == "":
                missing.append(field)
        if missing:
            issues.append(
                {
                    "step": idx,
                    "command": command,
                    "error": "Missing required fields.",
                    "fields": missing,
                }
            )

        if command in {"set-style", "format-range"}:
            try:
                parse_range_list(step.get("range"), step.get("ranges"))
            except Exception as exc:
                issues.append(
                    {
                        "step": idx,
                        "command": command,
                        "error": f"Could not parse ranges: {exc}",
                    }
                )
            for key in ("font", "fill", "border", "alignment", "theme"):
                if key in step and step.get(key) is not None:
                    try:
                        maybe_json_or_object(step.get(key))
                    except Exception as exc:
                        issues.append(
                            {
                                "step": idx,
                                "command": command,
                                "error": f"Malformed {key} payload: {exc}",
                            }
                        )
            if command == "set-style":
                if not any(
                    step.get(name)
                    for name in ("preset", "font", "fill", "border", "alignment", "number_format")
                ):
                    issues.append(
                        {
                            "step": idx,
                            "command": command,
                            "warning": "set-style has no style payload; add a preset or style fields to avoid a no-op.",
                        }
                    )
        if command == "copy-range":
            copy_flags = [
                bool(step.get("copy_values")),
                bool(step.get("copy_formulas")),
                bool(step.get("copy_styles")),
                bool(step.get("copy_merges")),
                bool(step.get("copy_dimensions")),
            ]
            if not any(copy_flags):
                issues.append(
                    {
                        "step": idx,
                        "command": command,
                        "warning": "No explicit copy flags supplied; defaulting to values, formulas, and styles.",
                    }
                )
        if command == "cleanup-sheet":
            style_targets = sum(
                len(parse_range_list(None, step.get(key)))
                for key in (
                    "title_ranges",
                    "section_header_ranges",
                    "table_header_ranges",
                    "input_ranges",
                    "formula_ranges",
                    "note_ranges",
                    "total_ranges",
                    "numeric_ranges",
                )
            )
            if style_targets == 0:
                issues.append(
                    {
                        "step": idx,
                        "command": command,
                        "warning": "cleanup-sheet has no explicit target ranges; it will only apply sheet-level polish.",
                    }
                )
    return issues


def namespace_from_action(step: dict[str, Any]) -> SimpleNamespace:
    normalized = normalize_plan_action(step)
    defaults = {
        "sheet": None,
        "save": False,
        "index": None,
        "amount": 1,
        "set": [],
        "font": None,
        "fill": None,
        "border": None,
        "alignment": None,
        "number_format": None,
        "horizontal": None,
        "vertical": None,
        "wrap_text": False,
        "style_name": "TableStyleMedium2",
        "show_first_column": False,
        "show_last_column": False,
        "show_column_stripes": False,
        "no_row_stripes": False,
        "copy_values": False,
        "copy_formulas": False,
        "copy_styles": False,
        "copy_merges": False,
        "copy_dimensions": False,
        "strict": False,
        "backup_policy": "managed",
        "range": None,
        "ranges": [],
        "preset": [],
        "theme": None,
        "show_gridlines": None,
        "tab_color": None,
        "zoom": None,
        "activate": False,
        "page_orientation": None,
        "fit_to_width": None,
        "fit_to_height": None,
        "paper_size": None,
        "center_horizontally": None,
        "center_vertically": None,
        "freeze_cell": None,
        "title_ranges": [],
        "section_header_ranges": [],
        "table_header_ranges": [],
        "input_ranges": [],
        "formula_ranges": [],
        "note_ranges": [],
        "total_ranges": [],
        "numeric_ranges": [],
        "label_ranges": [],
        "autofit_columns": [],
        "numeric_format": "number_comma_2dec",
    }
    merged = {**defaults, **normalized}
    return SimpleNamespace(**merged)


def ensure_managed_backup(
    workbook_path: Path,
    llm_work_root: Path,
    run_id: str,
    backup_policy: str,
    no_mirror_current: bool,
) -> tuple[str, dict[str, Any] | None]:
    backup_map_path = artifact_path(llm_work_root, run_id, "backup-map.json")
    if backup_policy == "skip":
        return "skipped", None
    if backup_map_path.exists():
        return "reused_existing", file_metadata(backup_map_path)
    if backup_policy != "managed":
        return "skipped", None

    command = [
        sys.executable,
        str(
            Path(__file__).resolve().parents[2]
            / "workbook-pipeline"
            / "backup-versioning"
            / "scripts"
            / "backup_workbooks.py"
        ),
        "--file",
        str(workbook_path),
        "--llm-work-root",
        str(llm_work_root),
        "--run-id",
        run_id,
        "--output",
        str(backup_map_path),
    ]
    if no_mirror_current:
        command.append("--no-mirror-current")
    completed = subprocess.run(command, capture_output=True, text=True)
    if completed.returncode != 0:
        message = completed.stderr.strip() or completed.stdout.strip() or "managed backup failed"
        raise ValueError(message)
    return "managed_by_skill", file_metadata(backup_map_path)


def cmd_check_workbook_status(args, wb, workbook_path: Path) -> dict[str, object]:
    return build_workbook_state(wb, workbook_path)


def cmd_create_sheet(args, wb, _workbook_path: Path) -> dict[str, object]:
    if args.sheet in wb.sheetnames:
        created = False
        ws = wb[args.sheet]
    else:
        index = len(wb.sheetnames) if args.index is None else max(0, min(args.index, len(wb.sheetnames)))
        ws = wb.create_sheet(title=args.sheet, index=index)
        created = True
    return {"sheet": ws.title, "created": created, "sheet_count": len(wb.sheetnames)}


def cmd_rename_sheet(args, wb, _workbook_path: Path) -> dict[str, object]:
    ws = wb[args.sheet]
    old_name = ws.title
    ws.title = args.new_name
    return {"old_name": old_name, "new_name": ws.title}


def cmd_reorder_sheet(args, wb, _workbook_path: Path) -> dict[str, object]:
    ws = wb[args.sheet]
    wb._sheets.remove(ws)
    index = max(0, min(args.index, len(wb._sheets)))
    wb._sheets.insert(index, ws)
    return {"sheet": ws.title, "new_index": index}


def cmd_delete_sheet(args, wb, _workbook_path: Path) -> dict[str, object]:
    if len(wb.sheetnames) == 1:
        raise ValueError("Cannot delete the only sheet in a workbook.")
    ws = wb[args.sheet]
    wb.remove(ws)
    return {"deleted_sheet": args.sheet, "remaining_sheets": wb.sheetnames}


def cmd_write_cells(args, wb, _workbook_path: Path) -> dict[str, object]:
    ws = workbook_sheet(wb, args.sheet)
    changes = []
    for raw in args.set:
        cell_ref, value_text = parse_assignment(raw)
        value = infer_value(value_text)
        ws[cell_ref] = value
        changes.append({"cell": cell_ref, "value": value})
    return {"sheet": ws.title, "change_count": len(changes), "changes": changes}


def cmd_write_formulas(args, wb, _workbook_path: Path) -> dict[str, object]:
    ws = workbook_sheet(wb, args.sheet)
    changes = []
    for raw in args.set:
        cell_ref, formula = parse_assignment(raw)
        if not formula.startswith("="):
            formula = "=" + formula
        ws[cell_ref] = formula
        changes.append({"cell": cell_ref, "formula": formula})
    return {"sheet": ws.title, "change_count": len(changes), "changes": changes}


def cmd_fill_formula(args, wb, _workbook_path: Path) -> dict[str, object]:
    ws = workbook_sheet(wb, args.sheet)
    source = ws[args.source_cell]
    if not source.value or not str(source.value).startswith("="):
        raise ValueError(f"Source cell {args.source_cell} does not contain a formula.")
    target_cells = cells_from_reference(ws, args.target_range)
    filled = 0
    for cell in target_cells:
        if cell.coordinate == source.coordinate:
            continue
        translator = Translator(source.value, origin=source.coordinate)
        cell.value = translator.translate_formula(
            row_delta=cell.row - source.row,
            col_delta=cell.column - source.column,
        )
        filled += 1
    return {
        "sheet": ws.title,
        "source_cell": args.source_cell,
        "target_range": args.target_range,
        "cells_filled": filled,
    }


def cmd_copy_range(args, wb, _workbook_path: Path) -> dict[str, object]:
    ws = workbook_sheet(wb, args.sheet)
    target = ws[args.target]
    start_row = target.row
    start_col = target.column
    copy_values = bool(args.copy_values)
    copy_formulas = bool(args.copy_formulas)
    copy_styles = bool(args.copy_styles)
    copy_merges = bool(args.copy_merges)
    copy_dimensions = bool(args.copy_dimensions)
    if not any((copy_values, copy_formulas, copy_styles, copy_merges, copy_dimensions)):
        copy_values = True
        copy_formulas = True
        copy_styles = True

    copied = 0
    values_written = 0
    formulas_written = 0
    styles_written = 0
    for r_offset, c_offset, cell in iter_range_cells(ws, args.source):
        dest = ws.cell(row=start_row + r_offset, column=start_col + c_offset)
        is_formula = isinstance(cell.value, str) and cell.value.startswith("=")
        if is_formula and copy_formulas:
            translator = Translator(cell.value, origin=cell.coordinate)
            dest.value = translator.translate_formula(
                row_delta=dest.row - cell.row,
                col_delta=dest.column - cell.column,
            )
            formulas_written += 1
        elif (not is_formula) and copy_values:
            dest.value = cell.value
            values_written += 1
        if copy_styles and cell.has_style:
            dest._style = copy(cell._style)
            styles_written += 1
        copied += 1

    merged_written = 0
    if copy_merges and ":" in args.source:
        source_min_col, source_min_row, _, _ = range_boundaries(args.source)
        for merged_range in merged_ranges_within(ws, args.source):
            min_col, min_row, max_col, max_row = range_boundaries(merged_range)
            translated = (
                f"{get_column_letter(start_col + (min_col - source_min_col))}{start_row + (min_row - source_min_row)}:"
                f"{get_column_letter(start_col + (max_col - source_min_col))}{start_row + (max_row - source_min_row)}"
            )
            ws.merge_cells(translated)
            merged_written += 1

    dimension_changes = {"column_widths": 0, "row_heights": 0}
    if copy_dimensions and ":" in args.source:
        source_min_col, source_min_row, source_max_col, source_max_row = range_boundaries(args.source)
        for src_col in range(source_min_col, source_max_col + 1):
            src_letter = get_column_letter(src_col)
            dest_letter = get_column_letter(start_col + (src_col - source_min_col))
            src_dim = ws.column_dimensions[src_letter]
            if src_dim.width is not None:
                ws.column_dimensions[dest_letter].width = src_dim.width
                dimension_changes["column_widths"] += 1
        for src_row in range(source_min_row, source_max_row + 1):
            dest_row = start_row + (src_row - source_min_row)
            src_dim = ws.row_dimensions[src_row]
            if src_dim.height is not None:
                ws.row_dimensions[dest_row].height = src_dim.height
                dimension_changes["row_heights"] += 1

    if getattr(args, "strict", False) and copied == 0:
        raise ValueError("copy-range strict mode failed because no cells were copied.")

    return {
        "sheet": ws.title,
        "source": args.source,
        "target": args.target,
        "cells_considered": copied,
        "values_written": values_written,
        "formulas_written": formulas_written,
        "styles_written": styles_written,
        "merged_ranges_written": merged_written,
        "dimension_changes": dimension_changes,
        "copy_values": copy_values,
        "copy_formulas": copy_formulas,
        "copy_styles": copy_styles,
        "copy_merges": copy_merges,
        "copy_dimensions": copy_dimensions,
    }


def cmd_define_name(args, wb, _workbook_path: Path) -> dict[str, object]:
    defined_name = DefinedName(name=args.name, attr_text=args.refers_to)
    try:
        wb.defined_names.pop(args.name, None)
    except Exception:
        pass
    wb.defined_names.add(defined_name)
    return {"name": args.name, "refers_to": args.refers_to}


def cmd_create_table(args, wb, _workbook_path: Path) -> dict[str, object]:
    ws = workbook_sheet(wb, args.sheet)
    table = Table(displayName=args.name, ref=args.range)
    if args.style_name:
        table.tableStyleInfo = TableStyleInfo(
            name=args.style_name,
            showFirstColumn=args.show_first_column,
            showLastColumn=args.show_last_column,
            showRowStripes=not args.no_row_stripes,
            showColumnStripes=args.show_column_stripes,
        )
    ws.add_table(table)
    return {
        "sheet": ws.title,
        "name": args.name,
        "range": args.range,
        "style_name": args.style_name,
    }


def cmd_insert_rows(args, wb, _workbook_path: Path) -> dict[str, object]:
    ws = workbook_sheet(wb, args.sheet)
    ws.insert_rows(idx=args.index, amount=args.amount)
    return {"sheet": ws.title, "index": args.index, "amount": args.amount}


def cmd_delete_rows(args, wb, _workbook_path: Path) -> dict[str, object]:
    ws = workbook_sheet(wb, args.sheet)
    ws.delete_rows(idx=args.index, amount=args.amount)
    return {"sheet": ws.title, "index": args.index, "amount": args.amount}


def cmd_insert_cols(args, wb, _workbook_path: Path) -> dict[str, object]:
    ws = workbook_sheet(wb, args.sheet)
    ws.insert_cols(idx=args.index, amount=args.amount)
    return {"sheet": ws.title, "index": args.index, "amount": args.amount}


def cmd_delete_cols(args, wb, _workbook_path: Path) -> dict[str, object]:
    ws = workbook_sheet(wb, args.sheet)
    ws.delete_cols(idx=args.index, amount=args.amount)
    return {"sheet": ws.title, "index": args.index, "amount": args.amount}


def cmd_set_column_widths(args, wb, _workbook_path: Path) -> dict[str, object]:
    ws = workbook_sheet(wb, args.sheet)
    changes = []
    for raw in args.set:
        target, width_text = parse_assignment(raw)
        width = float(width_text)
        if ":" in target:
            start, end = target.split(":", 1)
            start_idx = ws[start + "1"].column if start.isalpha() else ws[start].column
            end_idx = ws[end + "1"].column if end.isalpha() else ws[end].column
            for idx in range(start_idx, end_idx + 1):
                col = get_column_letter(idx)
                ws.column_dimensions[col].width = width
                changes.append({"column": col, "width": width})
        else:
            ws.column_dimensions[target].width = width
            changes.append({"column": target, "width": width})
    return {"sheet": ws.title, "change_count": len(changes), "changes": changes}


def cmd_set_row_heights(args, wb, _workbook_path: Path) -> dict[str, object]:
    ws = workbook_sheet(wb, args.sheet)
    changes = []
    for raw in args.set:
        target, height_text = parse_assignment(raw)
        height = float(height_text)
        if ":" in target:
            start, end = target.split(":", 1)
            for idx in range(int(start), int(end) + 1):
                ws.row_dimensions[idx].height = height
                changes.append({"row": idx, "height": height})
        else:
            idx = int(target)
            ws.row_dimensions[idx].height = height
            changes.append({"row": idx, "height": height})
    return {"sheet": ws.title, "change_count": len(changes), "changes": changes}


def cmd_set_style(args, wb, _workbook_path: Path) -> dict[str, object]:
    ws = workbook_sheet(wb, args.sheet)
    ranges = parse_range_list(args.range, getattr(args, "ranges", None))
    if not ranges:
        raise ValueError("set-style requires at least one target range.")
    cells = unique_cells_for_ranges(ws, ranges)
    spec, style_context = build_style_spec(args)
    changed = apply_style_to_cells(cells, spec)
    return {
        "sheet": ws.title,
        "ranges": ranges,
        "change_count": changed,
        "style": spec,
        "presets": style_context["presets"],
        "theme": style_context["theme"],
    }


def cmd_format_range(args, wb, _workbook_path: Path) -> dict[str, object]:
    ws = workbook_sheet(wb, args.sheet)
    ranges = parse_range_list(args.range, getattr(args, "ranges", None))
    if not ranges:
        raise ValueError("format-range requires at least one target range.")
    cells = unique_cells_for_ranges(ws, ranges)
    spec: dict[str, Any] = {}
    if args.bold is not None:
        spec.setdefault("font", {})["bold"] = args.bold
    if args.fill:
        spec["fill"] = {"color": args.fill}
    if args.number_format:
        spec["number_format"] = resolve_number_format(args.number_format)
    if args.horizontal or args.vertical or args.wrap_text:
        spec["alignment"] = {
            "horizontal": args.horizontal,
            "vertical": args.vertical,
            "wrap_text": args.wrap_text,
        }
    changed = apply_style_to_cells(cells, spec)
    return {
        "sheet": ws.title,
        "ranges": ranges,
        "change_count": changed,
        "style": spec,
    }


def cmd_freeze_panes(args, wb, _workbook_path: Path) -> dict[str, object]:
    ws = workbook_sheet(wb, args.sheet)
    ws.freeze_panes = args.cell
    return {"sheet": ws.title, "freeze_panes": args.cell}


def cmd_merge_cells(args, wb, _workbook_path: Path) -> dict[str, object]:
    ws = workbook_sheet(wb, args.sheet)
    ws.merge_cells(args.range)
    return {"sheet": ws.title, "merged_range": args.range}


def cmd_unmerge_cells(args, wb, _workbook_path: Path) -> dict[str, object]:
    ws = workbook_sheet(wb, args.sheet)
    ws.unmerge_cells(args.range)
    return {"sheet": ws.title, "unmerged_range": args.range}


def cmd_set_sheet_view(args, wb, _workbook_path: Path) -> dict[str, object]:
    ws = workbook_sheet(wb, args.sheet)
    changes: dict[str, Any] = {"sheet": ws.title}
    if args.show_gridlines is not None:
        ws.sheet_view.showGridLines = bool(args.show_gridlines)
        changes["show_gridlines"] = bool(args.show_gridlines)
    if args.tab_color:
        ws.sheet_properties.tabColor = normalize_color(args.tab_color)
        changes["tab_color"] = normalize_color(args.tab_color)
    if args.zoom is not None:
        ws.sheet_view.zoomScale = int(args.zoom)
        changes["zoom"] = int(args.zoom)
    if args.activate:
        for other in wb.worksheets:
            other.sheet_view.tabSelected = False
        ws.sheet_view.tabSelected = True
        wb.active = wb.sheetnames.index(ws.title)
        changes["active_sheet"] = ws.title
    if args.page_orientation:
        ws.page_setup.orientation = args.page_orientation
        changes["page_orientation"] = args.page_orientation
    if args.fit_to_width is not None:
        ws.page_setup.fitToWidth = int(args.fit_to_width)
        changes["fit_to_width"] = int(args.fit_to_width)
    if args.fit_to_height is not None:
        ws.page_setup.fitToHeight = int(args.fit_to_height)
        changes["fit_to_height"] = int(args.fit_to_height)
    if args.paper_size is not None:
        ws.page_setup.paperSize = int(args.paper_size)
        changes["paper_size"] = int(args.paper_size)
    if args.center_horizontally is not None:
        ws.print_options.horizontalCentered = bool(args.center_horizontally)
        changes["center_horizontally"] = bool(args.center_horizontally)
    if args.center_vertically is not None:
        ws.print_options.verticalCentered = bool(args.center_vertically)
        changes["center_vertically"] = bool(args.center_vertically)
    return changes


def autofit_columns(ws: Worksheet, column_ranges: list[str], max_width: float = 42.0) -> list[dict[str, Any]]:
    changes: list[dict[str, Any]] = []
    for item in column_ranges:
        if ":" in item:
            start, end = item.split(":", 1)
            start_idx = ws[start + "1"].column if start.isalpha() else ws[start].column
            end_idx = ws[end + "1"].column if end.isalpha() else ws[end].column
            indices = range(start_idx, end_idx + 1)
        else:
            indices = [ws[item + "1"].column if item.isalpha() else ws[item].column]
        for idx in indices:
            letter = get_column_letter(idx)
            max_len = 0
            for cell in ws[letter]:
                if cell.value is None:
                    continue
                max_len = max(max_len, len(str(cell.value)))
            width = min(max(8.0, float(max_len + 2)), max_width)
            ws.column_dimensions[letter].width = width
            changes.append({"column": letter, "width": width})
    return changes


def cmd_cleanup_sheet(args, wb, _workbook_path: Path) -> dict[str, object]:
    ws = workbook_sheet(wb, args.sheet)
    applied: dict[str, Any] = {"sheet": ws.title}
    view_args = SimpleNamespace(
        sheet=ws.title,
        show_gridlines=args.show_gridlines,
        tab_color=args.tab_color,
        zoom=args.zoom,
        activate=args.activate,
        page_orientation=args.page_orientation,
        fit_to_width=args.fit_to_width,
        fit_to_height=args.fit_to_height,
        paper_size=args.paper_size,
        center_horizontally=args.center_horizontally,
        center_vertically=args.center_vertically,
    )
    view_changes = cmd_set_sheet_view(view_args, wb, _workbook_path)
    if len(view_changes) > 1:
        applied["sheet_view"] = view_changes

    if args.freeze_cell:
        ws.freeze_panes = args.freeze_cell
        applied["freeze_panes"] = args.freeze_cell

    style_groups = [
        ("title_ranges", ["section_header"], None),
        ("section_header_ranges", ["section_header"], None),
        ("table_header_ranges", ["table_header"], None),
        ("input_ranges", ["input_cell"], None),
        ("formula_ranges", ["formula_cell"], None),
        ("note_ranges", ["note_cell"], None),
        ("total_ranges", ["final_total"], None),
    ]
    style_results = {}
    for attr_name, presets, number_format in style_groups:
        ranges = parse_range_list(None, getattr(args, attr_name, None))
        if not ranges:
            continue
        style_args = SimpleNamespace(
            sheet=ws.title,
            range=None,
            ranges=ranges,
            preset=presets,
            theme=args.theme,
            font=None,
            fill=None,
            border=None,
            alignment=None,
            number_format=number_format,
        )
        result = cmd_set_style(style_args, wb, _workbook_path)
        style_results[attr_name] = {
            "ranges": result["ranges"],
            "change_count": result["change_count"],
            "presets": result["presets"],
        }

    numeric_ranges = parse_range_list(None, getattr(args, "numeric_ranges", None))
    if numeric_ranges:
        number_args = SimpleNamespace(
            sheet=ws.title,
            range=None,
            ranges=numeric_ranges,
            preset=[],
            theme=args.theme,
            font=None,
            fill=None,
            border=None,
            alignment=json.dumps({"horizontal": "right"}),
            number_format=args.numeric_format,
        )
        result = cmd_set_style(number_args, wb, _workbook_path)
        style_results["numeric_ranges"] = {
            "ranges": result["ranges"],
            "change_count": result["change_count"],
            "number_format": args.numeric_format,
        }

    label_ranges = parse_range_list(None, getattr(args, "label_ranges", None))
    if label_ranges:
        label_args = SimpleNamespace(
            sheet=ws.title,
            range=None,
            ranges=label_ranges,
            preset=[],
            theme=args.theme,
            font=json.dumps({"bold": True}),
            fill=None,
            border=None,
            alignment=json.dumps({"horizontal": "left", "vertical": "center"}),
            number_format=None,
        )
        result = cmd_set_style(label_args, wb, _workbook_path)
        style_results["label_ranges"] = {
            "ranges": result["ranges"],
            "change_count": result["change_count"],
        }

    if style_results:
        applied["styles"] = style_results

    autofit_targets = parse_range_list(None, getattr(args, "autofit_columns", None))
    if autofit_targets:
        applied["autofit_columns"] = autofit_columns(ws, autofit_targets)

    return applied


def cmd_validate_plan(args, _wb, _workbook_path: Path) -> dict[str, object]:
    actions = load_plan(resolve_path(args.plan))
    issues = validate_plan_actions(actions)
    errors = [issue for issue in issues if "error" in issue]
    warnings = [issue for issue in issues if "warning" in issue]
    return {
        "plan": str(resolve_path(args.plan)),
        "step_count": len(actions),
        "valid": len(errors) == 0,
        "warning_count": len(warnings),
        "error_count": len(errors),
        "issues": issues,
        "suggestion": (
            "Plan can run as-is."
            if not issues
            else "Address errors first. Warnings are optional cleanup hints."
        ),
    }


def cmd_run_plan(args, wb, workbook_path: Path) -> dict[str, object]:
    actions = load_plan(resolve_path(args.plan))
    validation_issues = validate_plan_actions(actions)
    errors = [issue for issue in validation_issues if "error" in issue]
    if errors:
        raise ValueError(f"Plan validation failed: {json.dumps(errors, indent=2)}")
    results: list[dict[str, Any]] = []
    for idx, step in enumerate(actions, start=1):
        command = step["command"]
        step_args = namespace_from_action(step)
        result = COMMAND_MAP[command](step_args, wb, workbook_path)
        results.append({"step": idx, "command": command, "result": result})
    return {
        "plan": str(resolve_path(args.plan)),
        "step_count": len(results),
        "validation_issues": validation_issues,
        "results": results,
    }


def compact_executor_summary(command: str, result: dict[str, Any]) -> dict[str, Any]:
    if command == "run-plan":
        results = result.get("results", [])
        commands = [step.get("command") for step in results[:10]]
        return {
            "plan": result.get("plan"),
            "step_count": result.get("step_count"),
            "validation_issue_count": len(result.get("validation_issues", [])),
            "commands_preview": commands,
        }
    if command == "validate-plan":
        return {
            "plan": result.get("plan"),
            "step_count": result.get("step_count"),
            "valid": result.get("valid"),
            "warning_count": result.get("warning_count"),
            "error_count": result.get("error_count"),
        }
    if command in {
        "write-cells",
        "write-formulas",
        "set-column-widths",
        "set-row-heights",
        "set-style",
        "format-range",
    }:
        return {
            "sheet": result.get("sheet"),
            "change_count": result.get("change_count"),
            "ranges": result.get("ranges") or result.get("range"),
        }
    if command == "copy-range":
        return {
            "sheet": result.get("sheet"),
            "source": result.get("source"),
            "target": result.get("target"),
            "values_written": result.get("values_written"),
            "formulas_written": result.get("formulas_written"),
            "styles_written": result.get("styles_written"),
            "merged_ranges_written": result.get("merged_ranges_written"),
        }
    if command == "cleanup-sheet":
        return {
            "sheet": result.get("sheet"),
            "freeze_panes": result.get("freeze_panes"),
            "style_groups": sorted((result.get("styles") or {}).keys()),
            "autofit_columns_count": len(result.get("autofit_columns") or []),
        }
    if command == "set-sheet-view":
        return result
    return result


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Run deterministic workbook actions.")
    parser.add_argument("--workbook", required=True, help="Workbook path.")
    parser.add_argument("--save", action="store_true", help="Save the workbook after applying the action.")
    parser.add_argument("--llm-work-root", help="Optional workbook-local llm_work directory.")
    parser.add_argument("--run-id", help="Optional run id for action artifacts.")
    parser.add_argument(
        "--backup-policy",
        choices=("managed", "skip"),
        default="managed",
        help="How saved edit workflows handle backups. Defaults to managed.",
    )
    parser.add_argument(
        "--no-mirror-current",
        action="store_true",
        help="Do not update llm_work/current state files for this run.",
    )

    subparsers = parser.add_subparsers(dest="command", required=True)

    subparsers.add_parser("check-workbook-status")

    p = subparsers.add_parser("create-sheet")
    p.add_argument("--sheet", required=True)
    p.add_argument("--index", type=int)

    p = subparsers.add_parser("rename-sheet")
    p.add_argument("--sheet", required=True)
    p.add_argument("--new-name", required=True)

    p = subparsers.add_parser("reorder-sheet")
    p.add_argument("--sheet", required=True)
    p.add_argument("--index", required=True, type=int)

    p = subparsers.add_parser("delete-sheet")
    p.add_argument("--sheet", required=True)

    p = subparsers.add_parser("write-cells")
    p.add_argument("--sheet")
    p.add_argument("--set", action="append", default=[], required=True)

    p = subparsers.add_parser("write-formulas")
    p.add_argument("--sheet")
    p.add_argument("--set", action="append", default=[], required=True)

    p = subparsers.add_parser("fill-formula")
    p.add_argument("--sheet")
    p.add_argument("--source-cell", required=True)
    p.add_argument("--target-range", required=True)

    p = subparsers.add_parser("copy-range")
    p.add_argument("--sheet")
    p.add_argument("--source", required=True)
    p.add_argument("--target", required=True)
    p.add_argument("--copy-values", action="store_true")
    p.add_argument("--copy-formulas", action="store_true")
    p.add_argument("--copy-styles", action="store_true")
    p.add_argument("--copy-merges", action="store_true")
    p.add_argument("--copy-dimensions", action="store_true")
    p.add_argument("--strict", action="store_true")

    p = subparsers.add_parser("define-name")
    p.add_argument("--name", required=True)
    p.add_argument("--refers-to", required=True)

    p = subparsers.add_parser("create-table")
    p.add_argument("--sheet")
    p.add_argument("--name", required=True)
    p.add_argument("--range", required=True)
    p.add_argument("--style-name", default="TableStyleMedium2")
    p.add_argument("--show-first-column", action="store_true")
    p.add_argument("--show-last-column", action="store_true")
    p.add_argument("--show-column-stripes", action="store_true")
    p.add_argument("--no-row-stripes", action="store_true")

    p = subparsers.add_parser("insert-rows")
    p.add_argument("--sheet")
    p.add_argument("--index", required=True, type=int)
    p.add_argument("--amount", type=int, default=1)

    p = subparsers.add_parser("delete-rows")
    p.add_argument("--sheet")
    p.add_argument("--index", required=True, type=int)
    p.add_argument("--amount", type=int, default=1)

    p = subparsers.add_parser("insert-cols")
    p.add_argument("--sheet")
    p.add_argument("--index", required=True, type=int)
    p.add_argument("--amount", type=int, default=1)

    p = subparsers.add_parser("delete-cols")
    p.add_argument("--sheet")
    p.add_argument("--index", required=True, type=int)
    p.add_argument("--amount", type=int, default=1)

    p = subparsers.add_parser("set-column-widths")
    p.add_argument("--sheet")
    p.add_argument("--set", action="append", default=[], required=True)

    p = subparsers.add_parser("set-row-heights")
    p.add_argument("--sheet")
    p.add_argument("--set", action="append", default=[], required=True)

    p = subparsers.add_parser("set-style")
    p.add_argument("--sheet")
    p.add_argument("--range", action="append", default=[], required=True)
    p.add_argument("--preset", action="append", default=[])
    p.add_argument("--theme", help='JSON like {"accent_fill":"1F4E78"}')
    p.add_argument("--font", help='JSON like {"bold": true, "size": 12, "color": "1F4E78"}')
    p.add_argument("--fill", help='JSON like {"color": "D9EAF7"}')
    p.add_argument("--border", help='JSON like {"style": "thin", "color": "9FBAD0", "sides": ["left","right","top","bottom"]}')
    p.add_argument("--alignment", help='JSON like {"horizontal": "center", "vertical": "center", "wrap_text": true}')
    p.add_argument("--number-format", help="Direct format string or preset like currency/accounting/percentage.")

    p = subparsers.add_parser("format-range")
    p.add_argument("--sheet")
    p.add_argument("--range", action="append", default=[], required=True)
    p.add_argument("--bold", action=argparse.BooleanOptionalAction, default=None)
    p.add_argument("--fill")
    p.add_argument("--number-format")
    p.add_argument(
        "--horizontal",
        choices=("left", "center", "right", "fill", "justify", "centerContinuous", "distributed"),
    )
    p.add_argument(
        "--vertical",
        choices=("top", "center", "bottom", "justify", "distributed"),
    )
    p.add_argument("--wrap-text", action="store_true")

    p = subparsers.add_parser("freeze-panes")
    p.add_argument("--sheet")
    p.add_argument("--cell", required=True)

    p = subparsers.add_parser("merge-cells")
    p.add_argument("--sheet")
    p.add_argument("--range", required=True)

    p = subparsers.add_parser("unmerge-cells")
    p.add_argument("--sheet")
    p.add_argument("--range", required=True)

    p = subparsers.add_parser("set-sheet-view")
    p.add_argument("--sheet", required=True)
    p.add_argument("--show-gridlines", action=argparse.BooleanOptionalAction, default=None)
    p.add_argument("--tab-color")
    p.add_argument("--zoom", type=int)
    p.add_argument("--activate", action="store_true")
    p.add_argument("--page-orientation", choices=("portrait", "landscape"))
    p.add_argument("--fit-to-width", type=int)
    p.add_argument("--fit-to-height", type=int)
    p.add_argument("--paper-size", type=int)
    p.add_argument("--center-horizontally", action=argparse.BooleanOptionalAction, default=None)
    p.add_argument("--center-vertically", action=argparse.BooleanOptionalAction, default=None)

    p = subparsers.add_parser("cleanup-sheet")
    p.add_argument("--sheet", required=True)
    p.add_argument("--theme", help='JSON like {"accent_fill":"1F4E78"}')
    p.add_argument("--title-range", action="append", dest="title_ranges", default=[])
    p.add_argument("--section-header-range", action="append", dest="section_header_ranges", default=[])
    p.add_argument("--table-header-range", action="append", dest="table_header_ranges", default=[])
    p.add_argument("--input-range", action="append", dest="input_ranges", default=[])
    p.add_argument("--formula-range", action="append", dest="formula_ranges", default=[])
    p.add_argument("--note-range", action="append", dest="note_ranges", default=[])
    p.add_argument("--total-range", action="append", dest="total_ranges", default=[])
    p.add_argument("--numeric-range", action="append", dest="numeric_ranges", default=[])
    p.add_argument("--label-range", action="append", dest="label_ranges", default=[])
    p.add_argument("--autofit-columns", action="append", default=[])
    p.add_argument("--freeze-cell")
    p.add_argument("--show-gridlines", action=argparse.BooleanOptionalAction, default=False)
    p.add_argument("--tab-color")
    p.add_argument("--zoom", type=int)
    p.add_argument("--activate", action="store_true")
    p.add_argument("--page-orientation", choices=("portrait", "landscape"))
    p.add_argument("--fit-to-width", type=int)
    p.add_argument("--fit-to-height", type=int)
    p.add_argument("--paper-size", type=int)
    p.add_argument("--center-horizontally", action=argparse.BooleanOptionalAction, default=None)
    p.add_argument("--center-vertically", action=argparse.BooleanOptionalAction, default=None)
    p.add_argument("--numeric-format", default="number_comma_2dec")

    p = subparsers.add_parser("run-plan")
    p.add_argument("--plan", required=True, help="Path to JSON plan with an actions list.")

    p = subparsers.add_parser("validate-plan")
    p.add_argument("--plan", required=True, help="Path to JSON plan with an actions list.")

    return parser


COMMAND_MAP = {
    "check-workbook-status": cmd_check_workbook_status,
    "create-sheet": cmd_create_sheet,
    "rename-sheet": cmd_rename_sheet,
    "reorder-sheet": cmd_reorder_sheet,
    "delete-sheet": cmd_delete_sheet,
    "write-cells": cmd_write_cells,
    "write-formulas": cmd_write_formulas,
    "fill-formula": cmd_fill_formula,
    "copy-range": cmd_copy_range,
    "define-name": cmd_define_name,
    "create-table": cmd_create_table,
    "insert-rows": cmd_insert_rows,
    "delete-rows": cmd_delete_rows,
    "insert-cols": cmd_insert_cols,
    "delete-cols": cmd_delete_cols,
    "set-column-widths": cmd_set_column_widths,
    "set-row-heights": cmd_set_row_heights,
    "set-style": cmd_set_style,
    "format-range": cmd_format_range,
    "freeze-panes": cmd_freeze_panes,
    "merge-cells": cmd_merge_cells,
    "unmerge-cells": cmd_unmerge_cells,
    "set-sheet-view": cmd_set_sheet_view,
    "cleanup-sheet": cmd_cleanup_sheet,
    "run-plan": cmd_run_plan,
    "validate-plan": cmd_validate_plan,
}


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    started_at = datetime.now(timezone.utc)
    workbook_path = resolve_path(args.workbook)
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")
    llm_work_root, run_id = infer_action_paths(workbook_path, args.llm_work_root, args.run_id)
    workbook_info = workbook_metadata(workbook_path)
    task_hash = task_fingerprint(None)

    action_fn = COMMAND_MAP[args.command]
    wb = None if args.command == "validate-plan" else load_workbook(workbook_path)
    result = action_fn(args, wb, workbook_path)
    backup_outcome = None
    backup_artifact = None

    if args.save and wb is not None:
        backup_outcome, backup_artifact = ensure_managed_backup(
            workbook_path,
            llm_work_root,
            run_id,
            args.backup_policy,
            args.no_mirror_current,
        )
        save_workbook_safely(wb, workbook_path)

    payload = {
        "timestamp_utc": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "command": args.command,
        "workbook": str(workbook_path),
        "saved": args.save,
        "result": result,
    }
    output_path, _current_path = save_action_artifact(
        llm_work_root, run_id, args.command, payload, args.no_mirror_current
    )
    completed_at = datetime.now(timezone.utc)
    artifacts = {"action_result": file_metadata(output_path)}
    if backup_artifact:
        artifacts["backup_map"] = backup_artifact
    run_log_path = append_run_event(
        llm_work_root,
        run_id,
        {
            "event_type": "workbook-action.completed",
            "timestamp_utc": completed_at.replace(microsecond=0).isoformat(),
            "started_at_utc": started_at.replace(microsecond=0).isoformat(),
            "duration_ms": int((completed_at - started_at).total_seconds() * 1000),
            "skill": "workbook-action-executor",
            "status": "completed",
            "command": args.command,
            "source_file": str(workbook_path),
            "saved": args.save,
            "backup_policy": args.backup_policy if args.save else None,
            "backup_outcome": backup_outcome,
            "task_fingerprint": task_hash,
            "workbook": workbook_info,
            "artifacts": artifacts,
            "summary": compact_executor_summary(args.command, result),
        },
        update_current=not args.no_mirror_current,
    )

    print(f"Workbook: {workbook_path}")
    print(f"Command: {args.command}")
    if args.save:
        print("Workbook saved: true")
    if backup_outcome:
        print(f"Backup outcome: {backup_outcome}")
    print(f"Action artifact: {output_path}")
    print(f"Wrote run log: {run_log_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
