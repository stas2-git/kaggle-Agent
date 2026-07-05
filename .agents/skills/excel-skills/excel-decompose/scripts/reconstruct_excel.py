#!/usr/bin/env python3

import argparse
import re
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Color
from openpyxl.workbook.defined_name import DefinedName, DefinedNameDict


SHEET_PREFIX = "SHEET: "


def split_meta_line(line: str) -> tuple[str, str]:
    key, value = line.split(":", 1)
    return key.strip(), value.strip()


def infer_value(raw: str):
    if raw == "":
        return ""
    if raw == "TRUE":
        return True
    if raw == "FALSE":
        return False
    if raw in {"#N/A", "#VALUE!", "#REF!", "#DIV/0!", "#NUM!", "#NAME?", "#NULL!", "#CALC!"}:
        return raw
    if re.fullmatch(r"-?\d+", raw):
        try:
            return int(raw)
        except ValueError:
            return raw
    if re.fullmatch(r"-?(?:\d+\.\d+|\d+\.|\.\d+)", raw):
        try:
            return float(raw)
        except ValueError:
            return raw
    if re.fullmatch(r"-?(?:\d+(?:\.\d*)?|\.\d+)[eE][+-]?\d+", raw):
        try:
            return float(raw)
        except ValueError:
            return raw
    return raw


def parse_defined_name(line: str) -> dict:
    body = line[2:]
    name_part, value = body.split(":", 1)
    value = value.strip()

    local_sheet_id = None
    match = re.match(r"(.+?) \[localSheetId=(\d+)\]$", name_part.strip())
    if match:
        name = match.group(1).strip()
        local_sheet_id = int(match.group(2))
    else:
        name = name_part.strip()

    return {"name": name, "local_sheet_id": local_sheet_id, "value": value}


def parse_cell_line(line: str) -> dict:
    body = line[4:]
    ref, remainder = body.split(": ", 1)

    if " | " in remainder:
        parts = remainder.split(" | ")
    else:
        parts = [remainder]

    formula = ""
    value = ""
    for part in parts:
        if part.startswith("formula="):
            formula = part[len("formula="):]
        elif part.startswith("value="):
            value = part[len("value="):]
    return {"ref": ref.strip(), "formula": formula, "value": value}


def parse_decomposition(text: str) -> dict:
    lines = text.splitlines()
    idx = 0
    data = {"defined_names": [], "sheets": []}

    while idx < len(lines):
        line = lines[idx]

        if line == "DEFINED NAMES":
            idx += 1
            while idx < len(lines) and not lines[idx].startswith(SHEET_PREFIX):
                current = lines[idx]
                if current.startswith("- ") and current != "- none":
                    data["defined_names"].append(parse_defined_name(current))
                idx += 1
            continue

        if line.startswith(SHEET_PREFIX):
            sheet = {
                "name": line[len(SHEET_PREFIX):],
                "merged_ranges": [],
                "column_widths": [],
                "hidden_columns": [],
                "hidden_rows": [],
                "cells": [],
            }
            idx += 1
            in_cells = False

            while idx < len(lines):
                current = lines[idx]
                if current.startswith(SHEET_PREFIX):
                    break
                if current == "":
                    idx += 1
                    continue
                if current == "cells:":
                    in_cells = True
                    idx += 1
                    continue
                if in_cells and current.startswith("  - "):
                    if current == "  - none":
                        idx += 1
                        continue
                    sheet["cells"].append(parse_cell_line(current))
                    idx += 1
                    continue
                if current == "merged_ranges:":
                    idx += 1
                    while idx < len(lines) and lines[idx].startswith("  - "):
                        sheet["merged_ranges"].append(lines[idx][4:].strip())
                        idx += 1
                    continue
                if current == "column_widths:":
                    idx += 1
                    while idx < len(lines) and lines[idx].startswith("  - "):
                        column_name, width = lines[idx][4:].split(":", 1)
                        sheet["column_widths"].append((column_name.strip(), width.strip()))
                        idx += 1
                    continue
                if current == "hidden_columns:":
                    idx += 1
                    while idx < len(lines) and lines[idx].startswith("  - "):
                        sheet["hidden_columns"].append(lines[idx][4:].strip())
                        idx += 1
                    continue
                if current == "hidden_rows:":
                    idx += 1
                    while idx < len(lines) and lines[idx].startswith("  - "):
                        sheet["hidden_rows"].append(lines[idx][4:].strip())
                        idx += 1
                    continue
                if ":" in current:
                    key, value = split_meta_line(current)
                    sheet[key] = value
                idx += 1

            data["sheets"].append(sheet)
            continue

        idx += 1

    return data


def infer_number_format(raw: str) -> Optional[str]:
    raw = raw.strip()
    if raw == "":
        return None
    if re.fullmatch(r"-?(?:\d+(?:\.\d*)?|\.\d+)[eE][+-]?\d+", raw):
        return "0.00E+00"

    numeric_value = infer_value(raw)
    if not isinstance(numeric_value, (int, float)):
        return None

    if abs(float(numeric_value)) >= 1000:
        return "#,##0"
    return "0.00"


def build_workbook(parsed: dict, macro_shell: Optional[Path] = None) -> Workbook:
    if macro_shell is None:
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)
    else:
        wb = load_workbook(macro_shell, keep_vba=True)
        for ws in list(wb.worksheets):
            wb.remove(ws)

    for sheet_info in parsed["sheets"]:
        ws = wb.create_sheet(title=sheet_info["name"][:31] or "Sheet")

        if sheet_info.get("freeze_panes"):
            ws.freeze_panes = sheet_info["freeze_panes"]
        if sheet_info.get("tab_color"):
            ws.sheet_properties.tabColor = Color(rgb=sheet_info["tab_color"])
        if sheet_info.get("auto_filter"):
            ws.auto_filter.ref = sheet_info["auto_filter"]

        for merged in sheet_info.get("merged_ranges", []):
            if merged:
                try:
                    ws.merge_cells(merged)
                except ValueError:
                    pass

        for column_name, width in sheet_info.get("column_widths", []):
            try:
                ws.column_dimensions[column_name].width = float(width)
            except ValueError:
                continue

        for column_name in sheet_info.get("hidden_columns", []):
            ws.column_dimensions[column_name].hidden = True

        for row_number in sheet_info.get("hidden_rows", []):
            try:
                ws.row_dimensions[int(row_number)].hidden = True
            except ValueError:
                continue

        for cell in sheet_info.get("cells", []):
            target = ws[cell["ref"]]
            if cell["formula"]:
                formula = cell["formula"]
                if not formula.startswith("="):
                    formula = "=" + formula
                target.value = formula
            else:
                target.value = infer_value(cell["value"])

            number_format = infer_number_format(cell["value"])
            if number_format:
                target.number_format = number_format

    wb.defined_names = DefinedNameDict()
    for dn in parsed["defined_names"]:
        if not dn["name"] or dn["value"] in {"#REF!", "#NAME?"}:
            continue
        try:
            defined_name = DefinedName(
                name=dn["name"],
                attr_text=dn["value"],
                localSheetId=dn["local_sheet_id"],
            )
            wb.defined_names.add(defined_name)
        except Exception:
            continue

    return wb


def main() -> int:
    parser = argparse.ArgumentParser(description="Reconstruct a workbook from a decomposition text file.")
    parser.add_argument("--input", required=True, help="Path to the decomposition text file")
    parser.add_argument("--output", required=True, help="Path to the rebuilt .xlsx workbook")
    parser.add_argument(
        "--macro-shell",
        help="Optional .xlsm workbook to use as a macro-preserving shell for reconstruction",
    )
    args = parser.parse_args()

    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()
    macro_shell = Path(args.macro_shell).expanduser().resolve() if args.macro_shell else None

    if not input_path.exists():
        raise SystemExit(f"Input file not found: {input_path}")
    if macro_shell is not None and not macro_shell.exists():
        raise SystemExit(f"Macro shell not found: {macro_shell}")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    parsed = parse_decomposition(input_path.read_text(encoding="utf-8"))
    workbook = build_workbook(parsed, macro_shell=macro_shell)
    workbook.save(output_path)
    print(output_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
