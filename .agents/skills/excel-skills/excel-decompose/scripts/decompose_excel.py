#!/usr/bin/env python3

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from _shared.llm_work_audit import (  # type: ignore  # noqa: E402
    append_run_event,
    artifact_path,
    can_reuse_for_workbook,
    file_metadata,
    infer_llm_work_root,
    task_fingerprint,
    timestamp_run_id,
    workbook_metadata,
)

VALID_INCLUDES = {"backup", "summary", "plan", "checklist"}


def normalize_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value).replace("\n", " ").replace("\r", " ").strip()


def sanitize_artifact_name(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "-", name).strip("-")
    return cleaned or "workbook"


def normalize_includes(raw_values: list[str]) -> list[str]:
    includes: list[str] = []
    for raw in raw_values:
        for item in raw.split(","):
            name = item.strip().lower()
            if not name:
                continue
            if name not in VALID_INCLUDES:
                raise ValueError(
                    f"Unsupported include '{name}'. Valid options: {', '.join(sorted(VALID_INCLUDES))}"
                )
            if name not in includes:
                includes.append(name)
    return includes


def run_helper(command: list[str]) -> str:
    completed = subprocess.run(command, capture_output=True, text=True)
    if completed.returncode != 0:
        message = completed.stderr.strip() or completed.stdout.strip() or "helper command failed"
        raise RuntimeError(message)
    return completed.stdout.strip()


def render_output(workbook_path: Path, wb_formula, wb_values) -> str:
    lines = []
    timestamp = dt.datetime.now().isoformat(timespec="seconds")

    lines.append("WORKBOOK DECOMPOSITION")
    lines.append(f"source_file: {workbook_path}")
    lines.append(f"generated_at: {timestamp}")
    lines.append("")

    lines.append("WORKBOOK SUMMARY")
    lines.append(f"sheet_count: {len(wb_formula.sheetnames)}")
    lines.append("sheet_order:")
    for idx, sheet_name in enumerate(wb_formula.sheetnames, start=1):
        lines.append(f"  {idx}. {sheet_name}")
    lines.append("")

    lines.append("DEFINED NAMES")
    defined_names = list(wb_formula.defined_names.values())
    if defined_names:
        for dn in defined_names:
            local_note = ""
            if dn.localSheetId is not None:
                local_note = f" [localSheetId={dn.localSheetId}]"
            lines.append(f"- {dn.name}{local_note}: {dn.attr_text}")
    else:
        lines.append("- none")
    lines.append("")

    for sheet_name in wb_formula.sheetnames:
        ws_formula = wb_formula[sheet_name]
        ws_values = wb_values[sheet_name]

        lines.append(f"SHEET: {sheet_name}")
        lines.append(f"dimension: {ws_formula.calculate_dimension()}")
        lines.append(f"non_empty_cell_count: {sum(1 for row in ws_formula.iter_rows() for cell in row if cell.value is not None)}")
        if ws_formula.sheet_view.selection and ws_formula.freeze_panes:
            lines.append(f"freeze_panes: {ws_formula.freeze_panes}")
        if ws_formula.sheet_properties.tabColor and ws_formula.sheet_properties.tabColor.rgb:
            lines.append(f"tab_color: {ws_formula.sheet_properties.tabColor.rgb}")
        if ws_formula.auto_filter and ws_formula.auto_filter.ref:
            lines.append(f"auto_filter: {ws_formula.auto_filter.ref}")
        if ws_formula.merged_cells.ranges:
            lines.append("merged_ranges:")
            for merged in ws_formula.merged_cells.ranges:
                lines.append(f"  - {merged}")
        column_widths = []
        hidden_columns = []
        for key, dim in ws_formula.column_dimensions.items():
            if len(key) != 1 and not key.isalpha():
                continue
            if dim.width is not None:
                column_widths.append((key, dim.width))
            if dim.hidden:
                hidden_columns.append(key)
        if column_widths:
            lines.append("column_widths:")
            for column_name, width in column_widths:
                lines.append(f"  - {column_name}: {width}")
        if hidden_columns:
            lines.append("hidden_columns:")
            for column_name in hidden_columns:
                lines.append(f"  - {column_name}")

        hidden_rows = [str(idx) for idx, dim in ws_formula.row_dimensions.items() if dim.hidden]
        if hidden_rows:
            lines.append("hidden_rows:")
            for row_number in hidden_rows:
                lines.append(f"  - {row_number}")

        lines.append("cells:")
        found = False
        for row in ws_formula.iter_rows():
            for cell in row:
                formula_value = cell.value
                display_value = ws_values[cell.coordinate].value
                if formula_value is None and display_value is None:
                    continue

                found = True
                formula_text = ""
                if isinstance(formula_value, str) and formula_value.startswith("="):
                    formula_text = formula_value[1:].replace("\n", " ").replace("\r", " ").strip()

                value_text = normalize_value(display_value if display_value is not None else formula_value)

                if formula_text and value_text:
                    lines.append(f"  - {cell.coordinate}: formula={formula_text} | value={value_text}")
                elif formula_text:
                    lines.append(f"  - {cell.coordinate}: formula={formula_text}")
                else:
                    lines.append(f"  - {cell.coordinate}: value={value_text}")

        if not found:
            lines.append("  - none")
        lines.append("")

    return "\n".join(lines).rstrip() + "\n"


def build_summary(wb_formula, vba_modules: list[dict]) -> dict[str, int]:
    return {
        "sheet_count": len(wb_formula.sheetnames),
        "defined_name_count": len(list(wb_formula.defined_names.values())),
        "vba_module_count": len(vba_modules),
    }


def build_downstream_artifact_paths(llm_work_root: Path, run_id: str) -> dict[str, Path]:
    return {
        "decomposition": artifact_path(llm_work_root, run_id, "decomposition.txt"),
        "summary_markdown": artifact_path(llm_work_root, run_id, "summary.md"),
        "summary_json": artifact_path(llm_work_root, run_id, "summary.json"),
        "plan_markdown": artifact_path(llm_work_root, run_id, "plan.md"),
        "plan_json": artifact_path(llm_work_root, run_id, "plan.json"),
        "checklist_json": artifact_path(llm_work_root, run_id, "checklist.json"),
        "checklist_markdown": artifact_path(llm_work_root, run_id, "checklist.md"),
        "backup_map": artifact_path(llm_work_root, run_id, "backup-map.json"),
    }


def extract_vba_modules(workbook_path: Path) -> list[dict]:
    try:
        from oletools.olevba import VBA_Parser
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "VBA extraction for .xlsm files requires oletools. Install oletools or decompose an .xlsx workbook."
        ) from exc

    modules: list[dict] = []
    parser = VBA_Parser(str(workbook_path))
    try:
        if not parser.detect_vba_macros():
            return modules

        for (_, stream_path, module_name, code) in parser.extract_macros():
            clean_name = module_name or Path(stream_path).name or "Module"
            modules.append(
                {
                    "module_name": clean_name,
                    "stream_path": stream_path,
                    "code": code.replace("\r\n", "\n").replace("\r", "\n").rstrip(),
                }
            )
    finally:
        parser.close()

    return modules


def append_vba_output(base_text: str, vba_modules: list[dict]) -> str:
    if not vba_modules:
        return base_text

    lines = [base_text.rstrip(), "", "VBA MODULES"]
    for module in vba_modules:
        lines.append(f"MODULE: {module['module_name']}")
        lines.append(f"stream_path: {module['stream_path']}")
        lines.append("code:")
        if module["code"]:
            for line in module["code"].split("\n"):
                lines.append(f"  {line}")
        else:
            lines.append("  <empty>")
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def main() -> int:
    parser = argparse.ArgumentParser(description="Decompose an .xlsx/.xlsm workbook into LLM-readable text.")
    parser.add_argument("--workbook", required=True, help="Path to the workbook to decompose")
    parser.add_argument("--output", help="Optional custom export path for the decomposition text file.")
    parser.add_argument(
        "--include",
        action="append",
        default=[],
        help="Optional downstream artifacts to generate: backup, summary, plan, checklist. Can be repeated or comma-separated.",
    )
    parser.add_argument(
        "--task",
        help="User task text used by downstream planning/checklist helpers when included.",
    )
    parser.add_argument(
        "--llm-work-root",
        help="Optional workbook-local llm_work directory for audit artifacts. Defaults to <workbook folder>/llm_work.",
    )
    parser.add_argument(
        "--run-id",
        help="Optional run id for audit artifacts. Defaults to a New York timestamp.",
    )
    parser.add_argument(
        "--no-mirror-current",
        action="store_true",
        help="Do not update llm_work/current state files for this run.",
    )
    parser.add_argument(
        "--force-refresh",
        action="store_true",
        help="Regenerate artifacts even when current state indicates reusable outputs.",
    )
    args = parser.parse_args()
    started_at = datetime.now(timezone.utc)
    try:
        includes = normalize_includes(args.include)
    except ValueError as exc:
        raise SystemExit(str(exc))
    if any(name in includes for name in ("plan", "checklist")) and not args.task:
        raise SystemExit("--task is required when --include plan or --include checklist is used.")

    workbook_path = Path(args.workbook).expanduser().resolve()

    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")
    llm_work_root = infer_llm_work_root(workbook_path, args.llm_work_root)
    run_id = args.run_id or timestamp_run_id()
    downstream_paths = build_downstream_artifact_paths(llm_work_root, run_id)
    run_output_path = downstream_paths["decomposition"]
    output_path = Path(args.output).expanduser().resolve() if args.output else run_output_path
    workbook_info = workbook_metadata(workbook_path)
    task_hash = task_fingerprint(args.task)
    required_artifacts = ["decomposition"]
    if "backup" in includes:
        required_artifacts.append("backup_map")
    if any(name in includes for name in ("summary", "plan", "checklist")):
        required_artifacts.extend(["summary_markdown", "summary_json"])
    if "plan" in includes or "checklist" in includes:
        required_artifacts.extend(["plan_markdown", "plan_json"])
    if "checklist" in includes:
        required_artifacts.extend(["checklist_markdown", "checklist_json"])
    can_reuse, current_state = (False, {})
    if not args.force_refresh:
        can_reuse, current_state = can_reuse_for_workbook(
            llm_work_root,
            workbook_info,
            required_artifacts,
            task_hash=task_hash if ("plan" in includes or "checklist" in includes) else None,
        )
    if can_reuse:
        reused_artifacts = current_state.get("artifacts", {})
        decomposition_path = Path(reused_artifacts["decomposition"]["path"])
        explicit_export_path = None
        if output_path != run_output_path and output_path != decomposition_path:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            output_path.write_text(decomposition_path.read_text(encoding="utf-8"), encoding="utf-8")
            explicit_export_path = output_path
        completed_at = datetime.now(timezone.utc)
        run_log_path = append_run_event(
            llm_work_root,
            run_id,
            {
                "event_type": "excel-decompose.reused",
                "timestamp_utc": completed_at.replace(microsecond=0).isoformat(),
                "started_at_utc": started_at.replace(microsecond=0).isoformat(),
                "duration_ms": int((completed_at - started_at).total_seconds() * 1000),
                "skill": "excel-decompose",
                "status": "completed",
                "source_file": str(workbook_path),
                "requested_includes": includes,
                "task": args.task,
                "task_fingerprint": task_hash,
                "workbook": workbook_info,
                "artifacts": {
                    key: reused_artifacts.get(key) for key in required_artifacts if key in reused_artifacts
                } | {"explicit_export": file_metadata(explicit_export_path)},
                "summary": {
                    "reused_existing_artifact": True,
                    "requested_includes": includes,
                    "artifact_keys": required_artifacts,
                },
            },
            update_current=not args.no_mirror_current,
        )
        print(f"Reused current artifacts for workbook: {workbook_path}")
        if explicit_export_path:
            print(f"Wrote custom export: {explicit_export_path}")
        print(f"Wrote run log: {run_log_path}")
        return 0

    wb_formula = load_workbook(workbook_path, data_only=False, keep_vba=True)
    wb_values = load_workbook(workbook_path, data_only=True, keep_vba=True)
    vba_modules = extract_vba_modules(workbook_path)

    output_text = append_vba_output(render_output(workbook_path, wb_formula, wb_values), vba_modules)
    run_output_path.parent.mkdir(parents=True, exist_ok=True)
    run_output_path.write_text(output_text, encoding="utf-8")
    explicit_export_path = None
    if output_path != run_output_path:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(output_text, encoding="utf-8")
        explicit_export_path = output_path

    summary = build_summary(wb_formula, vba_modules)
    helper_outputs: dict[str, str] = {}

    try:
        if "backup" in includes:
            helper_outputs["backup"] = run_helper(
                [
                    sys.executable,
                    str(Path(__file__).resolve().parents[2] / "workbook-pipeline" / "backup-versioning" / "scripts" / "backup_workbooks.py"),
                    "--file",
                    str(workbook_path),
                    "--llm-work-root",
                    str(llm_work_root),
                    "--run-id",
                    run_id,
                    *([] if not args.no_mirror_current else ["--no-mirror-current"]),
                ]
            )

        if any(name in includes for name in ("summary", "plan", "checklist")):
            helper_outputs["summary"] = run_helper(
                [
                    sys.executable,
                    str(Path(__file__).resolve().parents[2] / "workbook-pipeline" / "semantic-summarizer" / "scripts" / "summarize_workbook.py"),
                    "--input",
                    str(run_output_path),
                    "--output",
                    str(downstream_paths["summary_markdown"]),
                    "--json-output",
                    str(downstream_paths["summary_json"]),
                    "--llm-work-root",
                    str(llm_work_root),
                    "--run-id",
                    run_id,
                    *([] if not args.no_mirror_current else ["--no-mirror-current"]),
                ]
            )

        if "plan" in includes or "checklist" in includes:
            helper_outputs["plan"] = run_helper(
                [
                    sys.executable,
                    str(Path(__file__).resolve().parents[2] / "workbook-pipeline" / "change-planner" / "scripts" / "plan_changes.py"),
                    "--brief",
                    str(downstream_paths["summary_json"]),
                    "--task",
                    str(args.task),
                    "--output",
                    str(downstream_paths["plan_markdown"]),
                    "--json-output",
                    str(downstream_paths["plan_json"]),
                    "--llm-work-root",
                    str(llm_work_root),
                    "--run-id",
                    run_id,
                    *([] if not args.no_mirror_current else ["--no-mirror-current"]),
                ]
            )

        if "checklist" in includes:
            helper_outputs["checklist"] = run_helper(
                [
                    sys.executable,
                    str(Path(__file__).resolve().parents[2] / "workbook-pipeline" / "validation-checklist-builder" / "scripts" / "build_checklist.py"),
                    "--plan",
                    str(downstream_paths["plan_json"]),
                    "--output",
                    str(downstream_paths["checklist_json"]),
                    "--markdown-output",
                    str(downstream_paths["checklist_markdown"]),
                    "--llm-work-root",
                    str(llm_work_root),
                    "--run-id",
                    run_id,
                    *([] if not args.no_mirror_current else ["--no-mirror-current"]),
                ]
            )
    except Exception as exc:
        failed_at = datetime.now(timezone.utc)
        run_log_path = append_run_event(
            llm_work_root,
            run_id,
            {
                "event_type": "excel-decompose.failed",
                "timestamp_utc": failed_at.replace(microsecond=0).isoformat(),
                "started_at_utc": started_at.replace(microsecond=0).isoformat(),
                "duration_ms": int((failed_at - started_at).total_seconds() * 1000),
                "skill": "excel-decompose",
                "status": "failed",
                "source_file": str(workbook_path),
                "requested_includes": includes,
                "task": args.task,
                "task_fingerprint": task_hash,
                "workbook": workbook_info,
                "error": str(exc),
            },
            update_current=not args.no_mirror_current,
        )
        print(f"Wrote decomposition artifact: {run_output_path}")
        if explicit_export_path:
            print(f"Wrote custom export: {explicit_export_path}")
        print(f"Wrote run log: {run_log_path}")
        raise SystemExit(str(exc))

    artifacts: dict[str, object] = {
        "decomposition": file_metadata(run_output_path),
        "explicit_export": file_metadata(explicit_export_path),
    }
    if "backup" in includes:
        artifacts["backup_map"] = file_metadata(downstream_paths["backup_map"])
    if "summary" in includes or "plan" in includes or "checklist" in includes:
        artifacts["summary_markdown"] = file_metadata(downstream_paths["summary_markdown"])
        artifacts["summary_json"] = file_metadata(downstream_paths["summary_json"])
    if "plan" in includes or "checklist" in includes:
        artifacts["plan_markdown"] = file_metadata(downstream_paths["plan_markdown"])
        artifacts["plan_json"] = file_metadata(downstream_paths["plan_json"])
    if "checklist" in includes:
        artifacts["checklist_json"] = file_metadata(downstream_paths["checklist_json"])
        artifacts["checklist_markdown"] = file_metadata(downstream_paths["checklist_markdown"])

    completed_at = datetime.now(timezone.utc)
    run_log_path = append_run_event(
        llm_work_root,
        run_id,
        {
            "event_type": "excel-decompose.completed",
            "timestamp_utc": completed_at.replace(microsecond=0).isoformat(),
            "started_at_utc": started_at.replace(microsecond=0).isoformat(),
            "duration_ms": int((completed_at - started_at).total_seconds() * 1000),
            "skill": "excel-decompose",
            "status": "completed",
            "source_file": str(workbook_path),
            "requested_includes": includes,
            "task": args.task,
            "task_fingerprint": task_hash,
            "workbook": workbook_info,
            "artifacts": artifacts,
            "summary": {
                **summary,
                "requested_includes": includes,
                "helper_count": len(helper_outputs),
                "helper_names": list(helper_outputs.keys()),
                "decomposition_bytes": run_output_path.stat().st_size,
            },
        },
        update_current=not args.no_mirror_current,
    )

    print(f"Wrote decomposition artifact: {run_output_path}")
    if explicit_export_path:
        print(f"Wrote custom export: {explicit_export_path}")
    if helper_outputs:
        print(f"Triggered helpers: {', '.join(helper_outputs.keys())}")
    print(f"Wrote run log: {run_log_path}")
    print(f"Sheets decomposed: {summary['sheet_count']}")
    print(f"Defined names: {summary['defined_name_count']}")
    print(f"VBA modules: {summary['vba_module_count']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
